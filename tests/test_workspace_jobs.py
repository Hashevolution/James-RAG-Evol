"""W8-A — workspace jobs backbone.

Three layers:
  1. core.workspace helpers — register / execute / list / count / get
     with owner-scope semantics.
  2. The three default handlers (excel_build / doc_combine /
     entity_export) produce real files in a tmp results dir. Each
     handler asks ``server_llmwiki.rag_engine.wiki_generator`` for
     the entity index — we monkey-patch that with a stub so the
     tests don't depend on the live wiki.
  3. HTTP endpoints — feature gates (workspace.run_jobs /
     workspace.view / admin.data), owner scoping (cross-owner 404),
     download 404 when output not yet written.
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault(
    "JAMES_JWT_SECRET",
    "test-secret-for-workspace-jobs-32chars",
)

from utils.console import ensure_utf8_console  # noqa: E402
ensure_utf8_console()


def _api_key() -> str:
    env_v = os.environ.get("JAMES_API_KEY")
    if env_v:
        return env_v.strip()
    env_path = Path(__file__).resolve().parent.parent / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8-sig").splitlines():
            if line.startswith("JAMES_API_KEY="):
                return line.split("=", 1)[1].strip()
    return ""


class _JobsFixture(unittest.TestCase):
    """Point ``core.workspace`` at a tmp DB + tmp results dir."""

    def setUp(self):
        from core import workspace as ws
        self._tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self._tmp_db.close()
        self._tmp_dir = tempfile.mkdtemp(prefix="ws_results_")
        self._saved_db  = ws._DB_PATH
        self._saved_dir = ws._RESULT_DIR
        ws._DB_PATH    = self._tmp_db.name
        ws._RESULT_DIR = self._tmp_dir
        ws._init_db()

    def tearDown(self):
        import shutil
        from core import workspace as ws
        ws._DB_PATH    = self._saved_db
        ws._RESULT_DIR = self._saved_dir
        Path(self._tmp_db.name).unlink(missing_ok=True)
        try:
            shutil.rmtree(self._tmp_dir)
        except Exception:
            pass


class HelperTests(_JobsFixture):
    def test_register_then_get_roundtrip(self):
        from core.workspace import register_job, get_job
        jid = register_job("excel_build", ["e1", "e2"], owner="alice")
        row = get_job(jid)
        self.assertEqual(row["status"], "pending")
        self.assertEqual(row["job_type"], "excel_build")
        self.assertEqual(row["owner"], "alice")
        self.assertEqual(row["input_refs"], ["e1", "e2"])

    def test_register_rejects_unknown_job_type(self):
        from core.workspace import register_job
        with self.assertRaises(ValueError):
            register_job("nope.fake", [], owner="alice")

    def test_register_rejects_non_list_input_refs(self):
        from core.workspace import register_job
        with self.assertRaises(ValueError):
            register_job("excel_build", "not-a-list", owner="alice")

    def test_owner_scope_get_returns_none_for_other(self):
        from core.workspace import register_job, get_job
        jid = register_job("excel_build", [], owner="alice")
        self.assertIsNone(get_job(jid, requester_username="bob"))
        self.assertIsNotNone(get_job(jid, requester_username="alice"))

    def test_list_owner_filter(self):
        from core.workspace import register_job, list_jobs, count_jobs
        register_job("excel_build", [], owner="alice")
        register_job("doc_combine", [], owner="alice")
        register_job("excel_build", [], owner="bob")
        self.assertEqual(count_jobs(owner="alice"), 2)
        self.assertEqual(count_jobs(owner="bob"), 1)
        self.assertEqual(count_jobs(), 3)
        owners = {r["owner"] for r in list_jobs(owner="alice")}
        self.assertEqual(owners, {"alice"})

    def test_list_status_filter(self):
        from core.workspace import register_job, count_jobs
        register_job("excel_build", [], owner="alice")
        register_job("excel_build", [], owner="alice")
        self.assertEqual(count_jobs(status="pending"), 2)
        self.assertEqual(count_jobs(status="done"), 0)
        self.assertEqual(count_jobs(status="garbage"), 0)


class _StubWiki:
    """Pretend to be the live ``rag_engine.wiki_generator``. The
    handlers pass entity_id_index values into Path() and then call
    _read_frontmatter; we keep the index value identical to the
    entity_id so str(Path(...)) round-trip works on any platform."""
    def __init__(self, entities):
        # entities: dict of entity_id → frontmatter dict
        self._fm = entities
        self.entity_id_index = {eid: eid for eid in entities}

    def _read_frontmatter(self, path):
        return self._fm.get(str(path))


def _install_stub_wiki(entities):
    """Monkey-patch the live engine for the duration of one test."""
    import server_llmwiki as srv
    class _StubEngine:
        wiki_generator = _StubWiki(entities)
    saved = getattr(srv, "rag_engine", None)
    srv.rag_engine = _StubEngine()
    return saved


def _restore_engine(saved):
    import server_llmwiki as srv
    if saved is None:
        try: del srv.rag_engine
        except AttributeError: pass
    else:
        srv.rag_engine = saved


class HandlerTests(_JobsFixture):
    def setUp(self):
        super().setUp()
        self._saved_engine = _install_stub_wiki({
            "ent_a": {"name": "Alpha", "entity_type": "concept",
                      "sensitivity": "public", "summary": "alpha summary"},
            "ent_b": {"name": "Beta",  "entity_type": "person",
                      "sensitivity": "internal", "summary": "beta summary"},
        })

    def tearDown(self):
        _restore_engine(self._saved_engine)
        super().tearDown()

    def test_excel_build_creates_xlsx_with_rows(self):
        # Sanity-check the stub install — fail loudly if the
        # monkey-patch was not propagated, so the "no rows" symptom
        # doesn't masquerade as a handler bug.
        import server_llmwiki as srv
        self.assertIsInstance(srv.rag_engine.wiki_generator, _StubWiki,
            "stub wiki was not installed on srv.rag_engine")
        self.assertEqual(
            sorted(srv.rag_engine.wiki_generator.entity_id_index.keys()),
            ["ent_a", "ent_b"],
        )

        from core import workspace as ws
        from openpyxl import load_workbook
        jid = ws.register_job("excel_build", ["ent_a", "ent_b", "missing-id"],
                              owner="alice")
        row = ws.execute_job(jid)
        self.assertEqual(row["status"], "done", row.get("error"))
        out_full = os.path.join(ws._RESULT_DIR, jid)
        files = [f for f in os.listdir(out_full) if f.endswith(".xlsx")]
        self.assertEqual(len(files), 1)
        wb = load_workbook(os.path.join(out_full, files[0]))
        sheet = wb["entities"]
        # header row + 2 rows for ent_a/ent_b
        rows = [r for r in sheet.iter_rows(values_only=True)]
        self.assertEqual(rows[0],
                         ("entity_id", "name", "type", "sensitivity", "summary"))
        names = {r[1] for r in rows[1:]}
        self.assertEqual(names, {"Alpha", "Beta"},
                         f"got rows: {rows}")
        self.assertIn("missing", wb.sheetnames)

    def test_doc_combine_concatenates_to_markdown(self):
        from core.workspace import register_job, execute_job, _RESULT_DIR
        jid = register_job("doc_combine", ["ent_a", "ent_b"], owner="alice")
        row = execute_job(jid)
        self.assertEqual(row["status"], "done", row.get("error"))
        # output file present + non-empty
        out_full = os.path.join(_RESULT_DIR, jid)
        files = [f for f in os.listdir(out_full) if f.endswith(".md")]
        self.assertEqual(len(files), 1)
        text = Path(os.path.join(out_full, files[0])).read_text(encoding="utf-8")
        # Stub wiki has no real file → handler writes "_(read error: ...)_"
        # That's acceptable — the test is that the file is created
        # and contains both entity ids as section markers.
        self.assertIn("ent_a", text)
        self.assertIn("ent_b", text)

    def test_entity_export_writes_json(self):
        from core.workspace import register_job, execute_job, _RESULT_DIR
        jid = register_job("entity_export", [], owner="alice")
        row = execute_job(jid)
        self.assertEqual(row["status"], "done", row.get("error"))
        out_full = os.path.join(_RESULT_DIR, jid)
        files = [f for f in os.listdir(out_full) if f.endswith(".json")]
        self.assertEqual(len(files), 1)
        data = json.loads(Path(os.path.join(out_full, files[0]))
                          .read_text(encoding="utf-8"))
        # Both stub entities surface (no category filter).
        self.assertEqual(data["count"], 2)
        names = {it["name"] for it in data["items"]}
        self.assertEqual(names, {"Alpha", "Beta"})

    def test_entity_export_filters_by_category(self):
        from core.workspace import register_job, execute_job, _RESULT_DIR
        jid = register_job("entity_export", ["concept"], owner="alice")
        row = execute_job(jid)
        out_full = os.path.join(_RESULT_DIR, jid)
        files = [f for f in os.listdir(out_full) if f.endswith(".json")]
        data = json.loads(Path(os.path.join(out_full, files[0]))
                          .read_text(encoding="utf-8"))
        self.assertEqual({it["entity_type"] for it in data["items"]},
                         {"concept"})

    def test_handler_failure_marks_row_failed(self):
        """Force an exception inside a handler and verify the row
        transitions to failed with an ``error`` set."""
        from core.workspace import register_job, execute_job, HANDLERS
        # Inject a broken handler temporarily.
        HANDLERS["broken"] = lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("boom"))
        try:
            jid = register_job("excel_build", [], owner="alice")
            # Force the job_type to 'broken' on the row.
            from core.workspace import _get_conn
            conn = _get_conn()
            conn.execute("UPDATE jobs SET job_type='broken' WHERE job_id = ?", (jid,))
            conn.commit()
            conn.close()
            row = execute_job(jid)
            self.assertEqual(row["status"], "failed")
            self.assertIn("boom", row.get("error") or "")
        finally:
            HANDLERS.pop("broken", None)


class EndpointTests(_JobsFixture):
    @classmethod
    def setUpClass(cls):
        cls._api_key = _api_key()

    def setUp(self):
        if not self._api_key:
            self.skipTest("JAMES_API_KEY missing")
        super().setUp()
        self._saved_engine = _install_stub_wiki({
            "ent_a": {"name": "A", "entity_type": "concept",
                      "sensitivity": "public"},
        })

    def tearDown(self):
        _restore_engine(self._saved_engine)
        super().tearDown()

    def _client(self):
        from fastapi.testclient import TestClient
        import server_llmwiki as srv
        return TestClient(srv.app)

    def _hdr(self, username: str, role: str):
        from core.auth import create_token
        return {"Authorization": f"Bearer {create_token(username, role)}"}

    # ── /jobs/run ─────────────────────────────────────────────────
    def test_run_requires_jwt(self):
        r = self._client().post(
            "/jobs/run",
            params={"api_key": self._api_key},
            json={"job_type": "excel_build", "input_refs": []},
        )
        # api_key is valid but no Bearer → role=employee, workspace.run_jobs
        # default includes employee → passes gate, but no JWT subject → 401.
        self.assertEqual(r.status_code, 401)

    def test_run_external_role_denied_by_default(self):
        # external is NOT in workspace.run_jobs default_allowed.
        r = self._client().post(
            "/jobs/run",
            params={"api_key": self._api_key},
            json={"job_type": "excel_build", "input_refs": []},
            headers=self._hdr("ext1", "external"),
        )
        self.assertEqual(r.status_code, 403)

    def test_run_employee_happy_path_done(self):
        r = self._client().post(
            "/jobs/run",
            params={"api_key": self._api_key},
            json={"job_type": "excel_build", "input_refs": ["ent_a"]},
            headers=self._hdr("alice", "employee"),
        )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["status"], "done")
        self.assertTrue(body["output_path"].endswith(".xlsx"))
        self.assertEqual(body["owner"], "alice")

    def test_run_unknown_job_type_returns_400(self):
        r = self._client().post(
            "/jobs/run",
            params={"api_key": self._api_key},
            json={"job_type": "nope", "input_refs": []},
            headers=self._hdr("alice", "employee"),
        )
        self.assertEqual(r.status_code, 400)

    # ── /jobs/list, /jobs/{id} ────────────────────────────────────
    def test_list_scoped_to_owner(self):
        # Two owners.
        from core.workspace import register_job
        register_job("excel_build", [], owner="alice")
        register_job("excel_build", [], owner="bob")
        r = self._client().get(
            "/jobs/list",
            params={"api_key": self._api_key},
            headers=self._hdr("alice", "employee"),
        )
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["total"], 1)
        self.assertEqual(body["items"][0]["owner"], "alice")

    def test_detail_cross_owner_returns_404(self):
        from core.workspace import register_job
        jid = register_job("excel_build", [], owner="alice")
        r = self._client().get(
            f"/jobs/{jid}",
            params={"api_key": self._api_key},
            headers=self._hdr("bob", "employee"),
        )
        self.assertEqual(r.status_code, 404)

    # ── /jobs/{id}/download ───────────────────────────────────────
    def test_download_404_when_output_missing(self):
        from core.workspace import register_job
        jid = register_job("excel_build", [], owner="alice")
        # pending → no output_path yet
        r = self._client().get(
            f"/jobs/{jid}/download",
            params={"api_key": self._api_key},
            headers=self._hdr("alice", "employee"),
        )
        self.assertEqual(r.status_code, 404)

    # ── /admin/jobs/list ──────────────────────────────────────────
    def test_admin_list_requires_admin_data(self):
        r = self._client().get(
            "/admin/jobs/list",
            params={"api_key": self._api_key},
            headers=self._hdr("emp1", "employee"),
        )
        self.assertEqual(r.status_code, 403)

    def test_admin_list_sees_all_owners(self):
        from core.workspace import register_job
        register_job("excel_build", [], owner="alice")
        register_job("excel_build", [], owner="bob")
        r = self._client().get(
            "/admin/jobs/list",
            params={"api_key": self._api_key},
            headers=self._hdr("test-admin", "admin"),
        )
        self.assertEqual(r.status_code, 200, r.text)
        self.assertEqual(r.json()["total"], 2)


if __name__ == "__main__":
    unittest.main()
